"""
Cruce de vales de comida x hora de salida del checador.

Dado un rango de fechas:
  1. Lee los vales de comida (rnd_reembolsos, concepto contiene 'COMIDA').
  2. Identifica al empleado por nombre_beneficiario (cascada: rnd_empleados -> checador).
  3. Busca su ultima SALIDA de ese dia en registros.
  4. Marca Cumple (salida >= 17:30) / No cumple / Revisar, y genera un Excel.

Uso:
  python cruce_comidas_salida.py 2026-07-01 2026-07-08
  (sin argumentos usa FECHA_INICIO / FECHA_FIN de abajo)
"""
import sys
import json
import unicodedata
from datetime import datetime, date, time
from urllib.request import Request, urlopen
from urllib.parse import quote

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== Config =====
SUPABASE_URL = 'https://uqncsqstpcynjxnjhrqu.supabase.co'
ANON_KEY = 'sb_publishable_bY6BY3wa5Xm2JCG2fy4F3g_fFgS5OsA'

# Rango por defecto si no se pasan argumentos de linea de comando
FECHA_INICIO = '2026-07-01'
FECHA_FIN = '2026-07-08'

# Regla de negocio: hora minima de salida para que el vale cuente
HORA_MINIMA_SALIDA = time(17, 30)

OUTPUT = r'C:/Users/USUARIO/Downloads/cruce_comidas_{inicio}_{fin}.xlsx'


# ===== Supabase =====
def supabase_get(path, params=None):
    qs = ''
    if params:
        qs = '?' + '&'.join(f'{k}={quote(str(v), safe="*.,()=:")}' for k, v in params.items())
    url = f'{SUPABASE_URL}/rest/v1/{path}{qs}'
    req = Request(url, headers={
        'apikey': ANON_KEY,
        'Authorization': f'Bearer {ANON_KEY}',
        'Accept': 'application/json',
    })
    with urlopen(req, timeout=60) as r:
        return json.loads(r.read().decode())


def supabase_get_all(path, params):
    """Trae todas las filas paginando de 1000 en 1000."""
    todos = []
    offset = 0
    while True:
        p = dict(params)
        p['limit'] = 1000
        p['offset'] = offset
        chunk = supabase_get(path, p)
        todos.extend(chunk)
        if len(chunk) < 1000:
            break
        offset += 1000
    return todos


# ===== Normalizacion =====
def norm_nombre(s):
    """Mayusculas, sin acentos, espacios colapsados. '' si es None."""
    if not s:
        return ''
    s = unicodedata.normalize('NFKD', str(s))
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    s = s.upper()
    return ' '.join(s.split())


def parse_fecha_vale(s):
    """Parsea 'YYYY-MM-DD'. Corrige año corrupto tipo 0026 -> 2026. None si invalida."""
    if not s:
        return None
    try:
        y, m, d = str(s)[:10].split('-')
        y = int(y)
        if y < 100:          # 0026 -> 2026
            y += 2000
        return date(y, int(m), int(d))
    except (ValueError, TypeError):
        return None


# ===== Carga de catalogos =====
def cargar_checador():
    """Empleados del checador. Devuelve:
       by_code: {codigo_normalizado -> emp}, by_name: {nombre_norm -> [emp,...]}."""
    emps = supabase_get('empleados', {
        'select': 'id,codigo_empleado,nombre,apellido,horario_id,activo'
    })
    by_code, by_name = {}, {}
    for e in emps:
        cod = e.get('codigo_empleado')
        if cod is not None:
            cod = str(cod).strip()
            by_code[cod] = e
            by_code[cod.zfill(4)] = e   # tolera codigos con y sin padding
        full1 = norm_nombre(f"{e.get('nombre','')} {e.get('apellido','')}")
        full2 = norm_nombre(f"{e.get('apellido','')} {e.get('nombre','')}")
        for k in {full1, full2}:
            if k:
                by_name.setdefault(k, []).append(e)
    return by_code, by_name


def cargar_rnd_empleados():
    """rnd_empleados. Devuelve {nombre_norm -> codigo}."""
    res = supabase_get('rnd_empleados', {'select': 'codigo,nombre'})
    by_name = {}
    for r in res:
        k = norm_nombre(r.get('nombre'))
        if k and r.get('codigo'):
            by_name.setdefault(k, str(r['codigo']).strip())
    return by_name


# ===== Matching en cascada =====
def _match_aproximado(nombre_norm, checador_by_name):
    """Devuelve emp si comparte >=2 tokens con un unico candidato dominante, si no None."""
    tokens = set(nombre_norm.split())
    if len(tokens) < 2:
        return None
    mejor, mejor_score, empatados = None, 0, 0
    for cand_norm, emps in checador_by_name.items():
        score = len(tokens & set(cand_norm.split()))
        if score > mejor_score:
            mejor, mejor_score, empatados = emps[0], score, 1
        elif score == mejor_score and score > 0:
            empatados += 1
    if mejor_score >= 2 and empatados == 1:
        return mejor
    return None


def identificar_empleado(nombre_benef, rnd_by_name, checador_by_code, checador_by_name):
    """Cascada: rnd_empleados -> checador por nombre -> aproximado -> sin_id.
       Devuelve {'emp': emp|None, 'metodo': 'rnd'|'checador'|'aprox'|'sin_id'}."""
    nn = norm_nombre(nombre_benef)
    if not nn:
        return {'emp': None, 'metodo': 'sin_id'}

    # 1) via rnd_empleados -> codigo -> checador
    codigo = rnd_by_name.get(nn)
    if codigo:
        emp = checador_by_code.get(codigo) or checador_by_code.get(codigo.zfill(4))
        if emp:
            return {'emp': emp, 'metodo': 'rnd'}

    # 2) directo por nombre completo en checador
    cands = checador_by_name.get(nn)
    if cands and len(cands) == 1:
        return {'emp': cands[0], 'metodo': 'checador'}

    # 3) aproximado por tokens
    emp = _match_aproximado(nn, checador_by_name)
    if emp:
        return {'emp': emp, 'metodo': 'aprox'}

    # 4) sin identificar
    return {'emp': None, 'metodo': 'sin_id'}


# ===== Salidas y clasificacion =====
def _parse_ts(s):
    """Parsea 'YYYY-MM-DD HH:MM:SS[.ffffff]' o con 'T'. None si falla."""
    if not s:
        return None
    txt = str(s).replace('T', ' ')[:19]
    try:
        return datetime.strptime(txt, '%Y-%m-%d %H:%M:%S')
    except ValueError:
        return None


def indexar_salidas(registros):
    """{(empleado_id, date) -> datetime de la ULTIMA salida de ese dia}."""
    idx = {}
    for r in registros:
        if r.get('tipo_registro') != 'SALIDA':
            continue
        ts = _parse_ts(r.get('fecha_hora'))
        if ts is None:
            continue
        key = (r.get('empleado_id'), ts.date())
        if key not in idx or ts > idx[key]:
            idx[key] = ts
    return idx


def clasificar(ultima_salida):
    """CUMPLE si ultima_salida >= HORA_MINIMA_SALIDA; NO_CUMPLE en otro caso o si None."""
    if ultima_salida is None:
        return {'estado': 'NO_CUMPLE', 'nota': 'Sin checada de salida ese dia'}
    if ultima_salida.time() >= HORA_MINIMA_SALIDA:
        return {'estado': 'CUMPLE', 'nota': ''}
    return {'estado': 'NO_CUMPLE', 'nota': 'Salio antes de 17:30'}


# ===== Carga de datos del rango =====
def cargar_vales_comida(fi, ff):
    """Vales con concepto que contiene 'comida' (case-insensitive), en el rango.
       Filtra por fecha ya corregida (año corrupto). Devuelve lista de dicts."""
    # 'comida' cubre COMIDAS/comida/Comida via ilike. Traemos amplio y filtramos por fecha en Python
    # para tolerar años corruptos (0026) que romperían un filtro de fecha en el servidor.
    res = supabase_get_all('rnd_reembolsos', {
        'select': 'fecha,nombre_beneficiario,monto,estado,concepto',
        'concepto': 'ilike.*comida*',
        'order': 'fecha.asc',
    })
    out = []
    for v in res:
        f = parse_fecha_vale(v.get('fecha'))
        if f is None or f < fi or f > ff:
            continue
        v['_fecha'] = f
        out.append(v)
    return out


def cargar_registros_salida(fi, ff):
    """Registros SALIDA del rango."""
    return supabase_get_all('registros', {
        'select': 'empleado_id,tipo_registro,fecha_hora',
        'tipo_registro': 'eq.SALIDA',
        'fecha_hora': f'gte.{fi}T00:00:00',
        'and': f'(fecha_hora.lte.{ff}T23:59:59)',
        'order': 'fecha_hora.asc',
    })


# ===== Construccion de filas del reporte =====
def _to_float(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0


def construir_filas(vales, rnd_by_name, checador_by_code, checador_by_name, salidas):
    """Una fila por vale. Detecta duplicados empleado-dia. Devuelve lista de dicts."""
    filas = []
    vistos = {}  # (emp_id, fecha) -> conteo, para marcar duplicados
    for v in vales:
        nombre_vale = v.get('nombre_beneficiario') or ''
        fecha = v.get('_fecha')
        monto = _to_float(v.get('monto'))
        ident = identificar_empleado(nombre_vale, rnd_by_name,
                                     checador_by_code, checador_by_name)
        emp = ident['emp']
        metodo = ident['metodo']

        if emp is None:
            filas.append({
                'fecha': fecha, 'nombre_vale': nombre_vale, 'codigo': '',
                'empleado_checador': '', 'monto': monto, 'estado_vale': v.get('estado', ''),
                'ultima_salida_str': '', 'cumple_str': '',
                'estado': 'REVISAR', 'nota': 'No se pudo identificar al empleado',
            })
            continue

        ultima = salidas.get((emp['id'], fecha))
        clas = clasificar(ultima)
        nombre_checador = norm_nombre(f"{emp.get('nombre','')} {emp.get('apellido','')}")

        # duplicado empleado-dia
        key = (emp['id'], fecha)
        vistos[key] = vistos.get(key, 0) + 1
        nota = clas['nota']
        estado = clas['estado']
        if vistos[key] > 1:
            estado = 'REVISAR'
            nota = (nota + ' | ' if nota else '') + 'Vale duplicado el mismo dia (revisar)'
        if metodo == 'aprox':
            estado = 'REVISAR' if estado != 'CUMPLE' else estado
            nota = (nota + ' | ' if nota else '') + 'Identificado por aproximacion (verificar)'

        filas.append({
            'fecha': fecha, 'nombre_vale': nombre_vale,
            'codigo': str(emp.get('codigo_empleado', '')),
            'empleado_checador': nombre_checador, 'monto': monto,
            'estado_vale': v.get('estado', ''),
            'ultima_salida_str': ultima.strftime('%Y-%m-%d %H:%M') if ultima else '',
            'cumple_str': 'SI' if clas['estado'] == 'CUMPLE' else 'NO',
            'estado': estado, 'nota': nota,
        })
    return filas


# ===== Excel =====
HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
OK_FILL = PatternFill('solid', fgColor='D5F5E3')
WARN_FILL = PatternFill('solid', fgColor='FCF3CF')
BAD_FILL = PatternFill('solid', fgColor='F5B7B1')
THIN = Side(border_style='thin', color='BDC3C7')
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
MONEY_FMT = '$#,##0.00'

FILL_POR_ESTADO = {'CUMPLE': OK_FILL, 'NO_CUMPLE': BAD_FILL, 'REVISAR': WARN_FILL}
ORDEN_ESTADO = {'CUMPLE': 0, 'NO_CUMPLE': 1, 'REVISAR': 2}

COLUMNAS = ['Fecha', 'Nombre en vale', 'Codigo', 'Empleado (checador)', 'Monto',
            'Estado vale', 'Ultima salida', 'Salio >= 5:30', 'Estado cruce', 'Notas']


def generar_excel(filas, fi, ff, ruta):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Comidas x Salida'

    # Titulo
    ws['A1'] = f'Cruce vales de comida x hora de salida  ({fi} a {ff})'
    ws['A1'].font = Font(bold=True, size=16, color='1F4E78')
    ws.append([])

    # Encabezados (fila 3)
    ws.append(COLUMNAS)
    hrow = ws.max_row
    for col in range(1, len(COLUMNAS) + 1):
        cell = ws.cell(row=hrow, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
    ws.freeze_panes = f'A{hrow + 1}'

    # Filas ordenadas por fecha y estado
    filas_ord = sorted(filas, key=lambda f: (f['fecha'] or date.min,
                                             ORDEN_ESTADO.get(f['estado'], 9)))
    for f in filas_ord:
        ws.append([
            f['fecha'].strftime('%Y-%m-%d') if f['fecha'] else '',
            f['nombre_vale'], f['codigo'], f['empleado_checador'], f['monto'],
            f['estado_vale'], f['ultima_salida_str'], f['cumple_str'],
            f['estado'], f['nota'],
        ])
        r = ws.max_row
        fill = FILL_POR_ESTADO.get(f['estado'])
        for col in range(1, len(COLUMNAS) + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = BORDER
            if fill:
                cell.fill = fill
        ws.cell(row=r, column=5).number_format = MONEY_FMT

    # Totales
    ws.append([])
    base = ws.max_row + 1
    def _tot(estado):
        sel = [f for f in filas if f['estado'] == estado]
        return len(sel), sum(f['monto'] for f in sel)
    for estado, etiqueta in [('CUMPLE', 'CUMPLEN (pagar)'),
                             ('NO_CUMPLE', 'NO CUMPLEN'),
                             ('REVISAR', 'A REVISAR')]:
        n, suma = _tot(estado)
        ws.cell(row=base, column=1, value=etiqueta).font = Font(bold=True)
        ws.cell(row=base, column=2, value=f'{n} vales').font = Font(bold=True)
        cel = ws.cell(row=base, column=5, value=suma)
        cel.number_format = MONEY_FMT
        cel.font = Font(bold=True)
        base += 1

    # Anchos
    for col, w in zip(range(1, len(COLUMNAS) + 1),
                      [12, 30, 10, 30, 12, 16, 18, 12, 14, 45]):
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(ruta)
    return ruta


# ===== Main =====
def main():
    fi_str = sys.argv[1] if len(sys.argv) > 1 else FECHA_INICIO
    ff_str = sys.argv[2] if len(sys.argv) > 2 else FECHA_FIN
    fi = parse_fecha_vale(fi_str)
    ff = parse_fecha_vale(ff_str)
    if fi is None or ff is None or fi > ff:
        print(f'Rango invalido: {fi_str} .. {ff_str}', file=sys.stderr)
        sys.exit(1)

    print(f'Rango: {fi} a {ff}')
    print('Cargando catalogos...')
    checador_by_code, checador_by_name = cargar_checador()
    rnd_by_name = cargar_rnd_empleados()
    print(f'  checador: {len(checador_by_code)} codigos | rnd: {len(rnd_by_name)} nombres')

    print('Cargando vales de comida...')
    vales = cargar_vales_comida(fi, ff)
    print(f'  {len(vales)} vales de comida en el rango')

    print('Cargando salidas...')
    registros = cargar_registros_salida(fi, ff)
    salidas = indexar_salidas(registros)
    print(f'  {len(registros)} checadas de salida | {len(salidas)} empleado-dia')

    filas = construir_filas(vales, rnd_by_name, checador_by_code,
                            checador_by_name, salidas)

    n_cumple = sum(1 for f in filas if f['estado'] == 'CUMPLE')
    n_no = sum(1 for f in filas if f['estado'] == 'NO_CUMPLE')
    n_rev = sum(1 for f in filas if f['estado'] == 'REVISAR')
    print(f'  Cumplen: {n_cumple} | No cumplen: {n_no} | Revisar: {n_rev}')

    ruta = OUTPUT.format(inicio=fi, fin=ff)
    generar_excel(filas, str(fi), str(ff), ruta)
    print(f'Excel generado: {ruta}')


if __name__ == '__main__':
    main()
