"""
Reporte de vales de comida que NO CUMPLEN, con cuanto tiempo nos deben.

Un vale "no cumple" cuando el empleado tiene vale de comida ese dia pero NO salio
a las 6:00 PM. Se calcula cuanto tiempo nos deben para llegar a la jornada:

  - Salio temprano (checo salida antes de 18:00):
        tiempo adeudado = 18:00 - hora de salida real.
  - Sin checada de salida (no hay evidencia de que trabajo):
        tiempo adeudado = 8.5 horas (jornada completa: 08:00-13:00 + 14:30-18:00).

Incluye sucursal y puesto del empleado. Solo lista los que NO cumplen.
Reutiliza la logica de identificacion de cruce_comidas_salida.py.

Uso:
  python reporte_no_cumplen.py 2026-01-01 2026-07-09
  (sin argumentos usa el rango de cruce_comidas_salida)
"""
import sys
from datetime import datetime, date, time, timedelta

import openpyxl
import openpyxl.worksheet.table
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import cruce_comidas_salida as base

# Hora de salida oficial de la jornada
HORA_SALIDA_OFICIAL = time(18, 0)
# Jornada neta cuando no hay checada (5h manana + 3.5h tarde)
JORNADA_COMPLETA_MIN = 8 * 60 + 30   # 510 minutos = 8.5 h


def minutos_a_horas(mins):
    """510 -> 8.5 (horas decimales, redondeado a 2)."""
    return round(mins / 60, 2)


def construir_filas_no_cumplen(vales, rnd_by_name, checador_by_code, checador_by_name, salidas):
    """Solo vales que NO cumplen. Calcula tiempo adeudado. Devuelve lista de dicts."""
    filas = []
    for v in vales:
        nombre_vale = v.get('nombre_beneficiario') or ''
        fecha = v.get('_fecha')
        monto = base._to_float(v.get('monto'))
        ident = base.identificar_empleado(nombre_vale, rnd_by_name,
                                          checador_by_code, checador_by_name)
        emp = ident['emp']
        if emp is None:
            continue  # sin identificar -> no entra en "no cumplen" (va a revisar en el otro reporte)

        ultima = salidas.get((emp['id'], fecha))

        # Determinar si cumple; si cumple, se omite
        if ultima is not None and ultima.time() >= base.HORA_MINIMA_SALIDA:
            continue  # CUMPLE -> fuera

        # No cumple: calcular tiempo adeudado
        if ultima is None:
            tiempo_min = JORNADA_COMPLETA_MIN
            motivo = 'Sin checada de salida (jornada completa)'
            salida_str = ''
        else:
            # 18:00 menos la hora de salida real, en minutos
            salida_dt = ultima
            oficial_dt = datetime.combine(salida_dt.date(), HORA_SALIDA_OFICIAL)
            delta = oficial_dt - salida_dt
            tiempo_min = max(0, int(delta.total_seconds() // 60))
            motivo = 'Salio antes de las 6:00 PM'
            salida_str = ultima.strftime('%H:%M')

        filas.append({
            'fecha': fecha,
            'sucursal': emp.get('sucursal') or '',
            'puesto': emp.get('puesto') or '',
            'codigo': str(emp.get('codigo_empleado', '')),
            'empleado': base.norm_nombre(f"{emp.get('nombre','')} {emp.get('apellido','')}"),
            'monto': monto,
            'salida_real': salida_str,
            'tiempo_min': tiempo_min,
            'tiempo_horas': minutos_a_horas(tiempo_min),
            'motivo': motivo,
        })
    return filas


# ===== Excel =====
HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
BAD_FILL = PatternFill('solid', fgColor='F5B7B1')
WARN_FILL = PatternFill('solid', fgColor='FCF3CF')
THIN = Side(border_style='thin', color='BDC3C7')
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
MONEY_FMT = '$#,##0.00'

DATE_FMT = 'yyyy-mm-dd'
HORAS_FMT = '0.00'

# Columnas planas, tipadas, listas para tabla dinamica (sin filas de resumen mezcladas)
COLUMNAS = ['Fecha', 'Sucursal', 'Puesto', 'Codigo', 'Empleado', 'Monto vale',
            'Salida real', 'Horas que deben', 'Motivo']


def generar_excel(filas, fi, ff, ruta):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'No cumplen'

    # Encabezados en la fila 1 (sin titulo arriba: una tabla limpia para dinamica)
    ws.append(COLUMNAS)
    for col in range(1, len(COLUMNAS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
    ws.freeze_panes = 'A2'

    # ordenar por sucursal, luego fecha
    filas_ord = sorted(filas, key=lambda f: (f['sucursal'], f['fecha'] or date.min))
    for f in filas_ord:
        ws.append([
            f['fecha'],                          # objeto date -> Excel lo trata como fecha
            f['sucursal'], f['puesto'], f['codigo'], f['empleado'],
            f['monto'],                          # numero
            f['salida_real'],                    # 'HH:MM' o '' (referencia, texto)
            f['tiempo_horas'],                   # numero decimal de horas
            f['motivo'],
        ])
        r = ws.max_row
        # amarillo si es "salio temprano", rojo si "sin checada"
        fill = WARN_FILL if f['salida_real'] else BAD_FILL
        for col in range(1, len(COLUMNAS) + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = BORDER
            cell.fill = fill
        ws.cell(row=r, column=1).number_format = DATE_FMT    # Fecha
        ws.cell(row=r, column=6).number_format = MONEY_FMT   # Monto
        ws.cell(row=r, column=8).number_format = HORAS_FMT   # Horas que deben

    # Formato de tabla real de Excel -> ideal para tablas dinamicas
    n_datos = len(filas_ord)
    if n_datos > 0:
        ref = f'A1:{get_column_letter(len(COLUMNAS))}{n_datos + 1}'
        tabla = openpyxl.worksheet.table.Table(displayName='NoCumplen', ref=ref)
        tabla.tableStyleInfo = openpyxl.worksheet.table.TableStyleInfo(
            name='TableStyleMedium2', showRowStripes=False)
        ws.add_table(tabla)

    for col, w in zip(range(1, len(COLUMNAS) + 1),
                      [12, 16, 22, 10, 30, 12, 12, 16, 34]):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ===== Hoja aparte: resumen por sucursal (numerico) =====
    ws2 = wb.create_sheet('Resumen sucursal')
    ws2.append(['Sucursal', 'Vales', 'Horas que deben', 'Monto'])
    for col in range(1, 5):
        c = ws2.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = BORDER
    por_suc = {}
    for f in filas:
        s = f['sucursal'] or '(sin sucursal)'
        d = por_suc.setdefault(s, {'n': 0, 'min': 0, 'monto': 0.0})
        d['n'] += 1
        d['min'] += f['tiempo_min']
        d['monto'] += f['monto']
    for s in sorted(por_suc, key=lambda k: -por_suc[k]['min']):
        d = por_suc[s]
        ws2.append([s, d['n'], minutos_a_horas(d['min']), d['monto']])
        r = ws2.max_row
        ws2.cell(row=r, column=3).number_format = HORAS_FMT
        ws2.cell(row=r, column=4).number_format = MONEY_FMT
    # fila total
    total_min = sum(f['tiempo_min'] for f in filas)
    total_monto = sum(f['monto'] for f in filas)
    ws2.append(['TOTAL', len(filas), minutos_a_horas(total_min), total_monto])
    r = ws2.max_row
    for col in range(1, 5):
        ws2.cell(row=r, column=col).font = Font(bold=True)
    ws2.cell(row=r, column=3).number_format = HORAS_FMT
    ws2.cell(row=r, column=4).number_format = MONEY_FMT
    for col, w in zip(range(1, 5), [18, 8, 16, 14]):
        ws2.column_dimensions[get_column_letter(col)].width = w

    wb.save(ruta)
    return ruta


def main():
    fi_str = sys.argv[1] if len(sys.argv) > 1 else base.FECHA_INICIO
    ff_str = sys.argv[2] if len(sys.argv) > 2 else base.FECHA_FIN
    fi = base.parse_fecha_vale(fi_str)
    ff = base.parse_fecha_vale(ff_str)
    if fi is None or ff is None or fi > ff:
        print(f'Rango invalido: {fi_str} .. {ff_str}', file=sys.stderr)
        sys.exit(1)

    print(f'Rango: {fi} a {ff}')
    print('Cargando catalogos...')
    checador_by_code, checador_by_name = base.cargar_checador()
    rnd_by_name = base.cargar_rnd_empleados()

    print('Cargando vales de comida...')
    vales = base.cargar_vales_comida(fi, ff)
    print(f'  {len(vales)} vales de comida en el rango')

    print('Cargando salidas...')
    registros = base.cargar_registros_salida(fi, ff)
    salidas = base.indexar_salidas(registros)

    filas = construir_filas_no_cumplen(vales, rnd_by_name, checador_by_code,
                                       checador_by_name, salidas)

    n_temprano = sum(1 for f in filas if f['salida_real'])
    n_sin = sum(1 for f in filas if not f['salida_real'])
    total_min = sum(f['tiempo_min'] for f in filas)
    print(f'  NO CUMPLEN: {len(filas)}  (temprano {n_temprano} | sin checada {n_sin})')
    print(f'  Tiempo total que deben: {total_min/60:.1f} horas')

    ruta = f'C:/Users/USUARIO/Downloads/no_cumplen_comidas_{fi}_{ff}.xlsx'
    generar_excel(filas, str(fi), str(ff), ruta)
    print(f'Excel generado: {ruta}')


if __name__ == '__main__':
    main()
